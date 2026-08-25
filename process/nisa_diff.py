"""つみたてNISA対象商品のExcelを読み、差分と保有照合をする"""

from __future__ import annotations

import io
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

USER_AGENT = "primary-info-bot/1.0 (news notifier; +https://github.com)"
SNAPSHOT_NAME = "nisa_snapshot.json"


def is_nisa_product_page(title: str, url: str) -> bool:
    t = title or ""
    u = (url or "").lower()
    return "nisa" in u or "つみたて投資枠対象商品" in t or ("つみたて" in t and "対象商品" in t)


def apply_nisa_diff(article: dict[str, Any], profile: dict[str, Any], data_dir: Path) -> bool:
    """成功したら easy_summary をセットして True。Excelが取れなければ False"""
    url = article.get("url", "")
    xlsx_url = _pick_xlsx_url(url)
    if not xlsx_url:
        return False

    funds = _download_fund_names(xlsx_url)
    if len(funds) < 20:
        return False

    snapshot_path = data_dir / SNAPSHOT_NAME
    previous = _load_snapshot(snapshot_path)
    prev_names = set(previous.get("funds", []))
    added = sorted(funds - prev_names)
    removed = sorted(prev_names - funds)
    first = not prev_names

    holdings = [str(h).strip() for h in profile.get("money", {}).get("holdings", []) if str(h).strip()]
    holding_hits, holding_missing = _match_holdings(holdings, funds)

    skip = False
    skip_reason = ""
    if not first and not added and not removed and not holding_missing:
        skip = True
        skip_reason = "商品リストに増減なし（保有分も対象のまま）"

    article["nisa_diff"] = {
        "count": len(funds),
        "added": added,
        "removed": removed,
        "first": first,
        "xlsx_url": xlsx_url,
        "holding_hits": holding_hits,
        "holding_missing": holding_missing,
    }
    article["quality"] = {"skip": skip, "skip_reason": skip_reason, "level": "product"}

    article["easy_summary"] = _build_summary(
        count=len(funds),
        added=added,
        removed=removed,
        first=first,
        holding_hits=holding_hits,
        holding_missing=holding_missing,
        xlsx_url=xlsx_url,
        skip=skip,
    )
    article.setdefault("decision_status", {})
    if not article["decision_status"]:
        from process.decision_status import classify_decision_status

        article["decision_status"] = classify_decision_status(article)

    _save_snapshot(snapshot_path, funds, xlsx_url)
    return True


def _pick_xlsx_url(page_url: str) -> str:
    try:
        response = requests.get(page_url, timeout=20, headers={"User-Agent": USER_AGENT})
        response.raise_for_status()
        response.encoding = response.apparent_encoding or "utf-8"
    except Exception:
        return ""

    soup = BeautifulSoup(response.text, "html.parser")
    company_list = ""
    asset_list = ""
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not re.search(r"\.xlsx?$", href, re.I):
            continue
        abs_url = urljoin(page_url, href)
        label = (a.get_text(" ", strip=True) + " " + href).lower()
        parent = a.find_parent("tr")
        row_text = parent.get_text(" ", strip=True) if parent else ""
        blob = label + " " + row_text
        if "運用会社" in blob or re.search(r"/23\.xlsx", href):
            company_list = abs_url
        elif "資産" in blob or re.search(r"/25\.xlsx", href):
            asset_list = abs_url
        elif not company_list:
            company_list = abs_url
    return company_list or asset_list


def _download_fund_names(xlsx_url: str) -> set[str]:
    response = requests.get(xlsx_url, timeout=40, headers={"User-Agent": USER_AGENT})
    response.raise_for_status()
    wb = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    names: set[str] = set()
    try:
        for sheet in wb.sheetnames:
            names.update(_names_from_sheet(wb[sheet]))
    finally:
        wb.close()
    return names


def _names_from_sheet(ws: Any) -> set[str]:
    names: set[str] = set()
    name_col: int | None = None
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if name_col is None:
            for idx, cell in enumerate(cells):
                if "ファンド名" in cell or cell == "ファンド":
                    name_col = idx
                    break
            continue
        if name_col >= len(cells):
            continue
        name = re.sub(r"\s+", " ", cells[name_col]).strip()
        if len(name) < 4:
            continue
        if any(bad in name for bad in ("ファンド名", "注", "本】", "対象商品")):
            continue
        names.add(name)
    return names


def _match_holdings(holdings: list[str], funds: set[str]) -> tuple[list[str], list[str]]:
    hits: list[str] = []
    missing: list[str] = []
    for holding in holdings:
        matches = [f for f in funds if _holding_matches(holding, f)]
        if not matches:
            missing.append(holding)
            continue
        preferred = [f for f in matches if "オール・カントリー" in f or "オールカントリー" in f]
        if "S&P500" in holding or "Ｓ＆Ｐ５００" in holding:
            sp = [f for f in matches if "S&P500" in f or "Ｓ＆Ｐ５００" in f]
            if sp:
                preferred = sp
        hits.append(preferred[0] if preferred else sorted(matches, key=len)[-1])
    return hits, missing


def _holding_matches(holding: str, fund: str) -> bool:
    h = re.sub(r"\s+", "", holding.lower())
    f = re.sub(r"\s+", "", fund.lower())
    if not h:
        return False
    return h in f or f in h


def _fmt_list(items: list[str], limit: int = 5) -> str:
    if not items:
        return "なし"
    shown = items[:limit]
    extra = len(items) - len(shown)
    text = " / ".join(shown)
    if extra > 0:
        text += f" ほか{extra}本"
    return text


def _build_summary(
    *,
    count: int,
    added: list[str],
    removed: list[str],
    first: bool,
    holding_hits: list[str],
    holding_missing: list[str],
    xlsx_url: str,
    skip: bool,
) -> dict[str, Any]:
    if holding_missing:
        relevance = "⚠️ 指定した保有ファンドが、今回の対象一覧から見つかりません。"
        action = "証券口座で積立設定を確認。見つからない商品は新規積立できなくなっている可能性あり。"
        impact = "対象外になると新規の積立は止まる。すでに持っている分は基本そのまま。"
    elif holding_hits:
        relevance = "指定した保有ファンドは、まだつみたて対象に入っている。"
        action = "今すぐ動く必要なし。次回の差分通知まで待ってOK。"
        impact = "いま積立中の商品は、この更新では対象から外れていない。"
    else:
        relevance = "つみたてNISA利用者向け。保有ファンド名を profile.yaml に書くと、自分の分だけ照合する。"
        action = "気になる商品が増減リストにあれば、口座で検索。"
        impact = ""

    if first:
        one_line = f"つみたてNISA対象は現在 {count} 本。今回から増減を監視する。"
        what_changed = f"初回スキャン完了（{count}本）。次回更新から「増えた/減った商品名」だけ通知する。"
        if holding_hits:
            what_changed += f" 保有照合OK: {_fmt_list(holding_hits, 3)}"
        if holding_missing:
            what_changed += f" 見つからない: {_fmt_list(holding_missing, 3)}"
        title = "NISA対象リストの監視を開始"
    elif not added and not removed:
        one_line = f"対象商品は {count} 本のまま。増減なし。"
        what_changed = "Excelは更新されているが、商品名の増減は検出されなかった。"
        title = "NISA対象リスト（増減なし）"
    else:
        one_line = f"対象 {count} 本。追加 {len(added)} / 除外 {len(removed)}。"
        parts = []
        if added:
            parts.append(f"追加: {_fmt_list(added)}")
        if removed:
            parts.append(f"除外: {_fmt_list(removed)}")
        what_changed = "　".join(parts)
        title = f"NISA対象 追加{len(added)} / 除外{len(removed)}"

    return {
        "easy_title": title,
        "one_line": one_line,
        "what_changed": what_changed,
        "gemini_relevance": relevance,
        "gemini_action": action,
        "gemini_impact": impact,
        "gemini_followup": "",
        "content_category": "money",
        "added_preview": _fmt_list(added) if added else "",
        "removed_preview": _fmt_list(removed) if removed else "",
        "holding_hits": holding_hits,
        "holding_missing": holding_missing,
        "xlsx_url": xlsx_url,
        "ai_used": False,
        "ai_status": "diff",
        "skip_notify": skip,
    }


def _load_snapshot(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_snapshot(path: Path, funds: set[str], xlsx_url: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "xlsx_url": xlsx_url,
        "count": len(funds),
        "funds": sorted(funds),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
